"""
MongoDB to Excel Exporter for Pecfest 2025
This script downloads all collections from MongoDB and exports them to Excel files.
It also creates separate Excel files for registrations grouped by event.
"""

import os
from datetime import datetime
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import urllib.parse
from typing import Dict, List, Any
import json

# MongoDB Connection String
# Update this with your MongoDB URI or set as environment variable
# MONGODB_URI = "mongodb+srv://pecfestdev_db_user:rfCA2MDfuvLd7rQB@cluster0.zlqaabw.mongodb.net/?appName=Cluster0"
MONGODB_URI="mongodb+srv://pecfestdev_db_user:rfCA2MDfuvLd7rQB@cluster0.zlqaabw.mongodb.net/test?appName=Cluster0"
DATABASE_NAME = "test"  # Change this to your actual database name if different

# Output directory for Excel files
OUTPUT_DIR = "output"
REGISTRATIONS_DIR = os.path.join(OUTPUT_DIR, "events")

# Collection names based on the models
COLLECTIONS = [
    "users",
    "events", 
    "registrations",
    "adminusers",
    "discounts",
    "minimarathons",
    "nonpecminimarathons",
    "registrationforms",
    "otps"
]

def create_output_directories():
    """Create output directories if they don't exist"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(REGISTRATIONS_DIR, exist_ok=True)
    print(f"✓ Output directories created: {OUTPUT_DIR}")

def connect_to_mongodb():
    """Connect to MongoDB and return client and database"""
    try:
        print("Connecting to MongoDB...")
        client = MongoClient(MONGODB_URI, serverSelectionTimeoutMS=5000)
        # Test connection
        client.server_info()
        
        # Get database - try from URI first, then use DATABASE_NAME
        try:
            db = client.get_default_database()
        except:
            db = client[DATABASE_NAME]
        
        print(f"✓ Connected to MongoDB: {db.name}")
        
        # List available collections
        collections = db.list_collection_names()
        print(f"✓ Found {len(collections)} collections in database")
        
        return client, db
    except Exception as e:
        print(f"✗ Failed to connect to MongoDB: {e}")
        raise

def flatten_document(doc: Dict, parent_key: str = '', separator: str = '_') -> Dict:
    """
    Flatten nested dictionaries and handle special types
    """
    items = []
    for key, value in doc.items():
        new_key = f"{parent_key}{separator}{key}" if parent_key else key
        
        # Handle ObjectId
        if hasattr(value, '__str__') and type(value).__name__ == 'ObjectId':
            items.append((new_key, str(value)))
        # Handle datetime
        elif isinstance(value, datetime):
            items.append((new_key, value.strftime('%Y-%m-%d %H:%M:%S')))
        # Handle nested dictionaries
        elif isinstance(value, dict):
            items.extend(flatten_document(value, new_key, separator=separator).items())
        # Handle lists
        elif isinstance(value, list):
            if len(value) > 0 and isinstance(value[0], dict):
                # For list of dicts, convert to JSON string
                items.append((new_key, json.dumps(value, default=str)))
            else:
                # For simple lists, join with comma
                items.append((new_key, ', '.join(map(str, value))))
        # Handle Buffer/Binary data (for posterImage in RegistrationForm)
        elif isinstance(value, bytes):
            items.append((new_key, f"<Binary data: {len(value)} bytes>"))
        # Handle None
        elif value is None:
            items.append((new_key, ''))
        else:
            items.append((new_key, value))
    
    return dict(items)

def style_header_row(ws, num_columns):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # Set width with some padding, max 50
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def export_collection_to_excel(db, collection_name: str, filename: str):
    """Export a MongoDB collection to Excel"""
    try:
        collection = db[collection_name]
        documents = list(collection.find())
        
        if not documents:
            print(f"  ⚠ Collection '{collection_name}' is empty, skipping...")
            return 0
        
        # Flatten all documents
        flattened_docs = [flatten_document(doc) for doc in documents]
        
        # Get all unique keys across all documents
        all_keys = set()
        for doc in flattened_docs:
            all_keys.update(doc.keys())
        
        # Sort keys for consistent column order
        headers = sorted(list(all_keys))
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = collection_name[:31]  # Excel sheet name limit
        
        # Write headers
        ws.append(headers)
        
        # Write data rows
        for doc in flattened_docs:
            row = [doc.get(key, '') for key in headers]
            ws.append(row)
        
        # Apply styling
        style_header_row(ws, len(headers))
        auto_adjust_column_width(ws)
        
        # Save workbook
        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
        
        print(f"  ✓ Exported {len(documents)} documents to {filename}")
        return len(documents)
        
    except Exception as e:
        print(f"  ✗ Error exporting collection '{collection_name}': {e}")
        return 0

def export_registrations_by_event(db):
    """Export registrations grouped by event to separate Excel files"""
    try:
        print("\nExporting event registrations...")
        
        # Get all events
        events_collection = db['events']
        events = list(events_collection.find())
        
        if not events:
            print("  ⚠ No events found in database")
            return
        
        print(f"  Found {len(events)} events")
        
        # Get all registrations
        registrations_collection = db['registrations']
        
        event_count = 0
        for event in events:
            event_id = event.get('eventId')
            event_name = event.get('eventName', 'Unknown Event')
            
            if not event_id:
                continue
            
            # Find all registrations for this event
            registrations = list(registrations_collection.find({'eventId': event_id}))
            
            if not registrations:
                print(f"  ⚠ No registrations for event: {event_name}")
                continue
            
            print(f"  📝 Processing: {event_name} (ID: {event_id})")
            
            # Flatten registrations
            flattened_regs = [flatten_document(reg) for reg in registrations]
            
            # Get all unique keys
            all_keys = set()
            for reg in flattened_regs:
                all_keys.update(reg.keys())
            
            headers = sorted(list(all_keys))
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            # Sanitize sheet title - remove invalid characters
            safe_sheet_name = "".join(c for c in event_name if c not in [':', '\\', '/', '?', '*', '[', ']'])[:31]
            ws.title = safe_sheet_name if safe_sheet_name else "Event"
            
            # Write headers
            ws.append(headers)
            
            # Write data
            for reg in flattened_regs:
                row = [reg.get(key, '') for key in headers]
                ws.append(row)
            
            # Apply styling
            style_header_row(ws, len(headers))
            auto_adjust_column_width(ws)
            
            # Create safe filename
            safe_event_name = "".join(c for c in event_name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_event_name = safe_event_name.replace(' ', '_')
            filename = f"{event_id}_{safe_event_name}_registrations.xlsx"
            filepath = os.path.join(REGISTRATIONS_DIR, filename)
            
            wb.save(filepath)
            event_count += 1
            print(f"  ✓ Exported {len(registrations)} registrations for: {event_name}")
        
        print(f"\n✓ Exported registrations for {event_count} events")
        return event_count
        
    except Exception as e:
        print(f"✗ Error exporting event registrations: {e}")
        import traceback
        traceback.print_exc()
        return 0

def export_all_collections(db):
    """Export all collections to Excel files"""
    print("\nExporting all collections...")
    
    total_documents = 0
    successful_exports = 0
    
    for collection_name in COLLECTIONS:
        try:
            filename = f"{collection_name}.xlsx"
            count = export_collection_to_excel(db, collection_name, filename)
            if count > 0:
                successful_exports += 1
                total_documents += count
        except Exception as e:
            print(f"  ✗ Failed to export {collection_name}: {e}")
    
    print(f"\n✓ Successfully exported {successful_exports} collections with {total_documents} total documents")
    return successful_exports, total_documents

def create_summary_report(db):
    """Create a summary report of all collections"""
    try:
        print("\nCreating summary report...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Database Summary"
        
        # Headers
        headers = ["Collection Name", "Document Count", "Export Status"]
        ws.append(headers)
        
        # Get collection stats
        for collection_name in COLLECTIONS:
            try:
                collection = db[collection_name]
                count = collection.count_documents({})
                status = "✓ Exported" if count > 0 else "⚠ Empty"
                ws.append([collection_name, count, status])
            except Exception as e:
                ws.append([collection_name, 0, f"✗ Error: {str(e)}"])
        
        # Add timestamp
        ws.append([])
        ws.append(["Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        
        # Apply styling
        style_header_row(ws, len(headers))
        auto_adjust_column_width(ws)
        
        # Save
        filepath = os.path.join(OUTPUT_DIR, "00_SUMMARY_REPORT.xlsx")
        wb.save(filepath)
        
        print(f"  ✓ Summary report created: 00_SUMMARY_REPORT.xlsx")
        
    except Exception as e:
        print(f"  ✗ Error creating summary report: {e}")

def main():
    """Main execution function"""
    print("=" * 70)
    print("MongoDB to Excel Exporter - Pecfest 2025")
    print("=" * 70)
    
    try:
        # Create directories
        create_output_directories()
        
        # Connect to MongoDB
        client, db = connect_to_mongodb()
        
        # Export all collections
        successful_collections, total_docs = export_all_collections(db)
        
        # Export registrations by event
        total_events = export_registrations_by_event(db)
        
        # Create summary report
        create_summary_report(db)
        
        # Close connection
        client.close()
        print("\n" + "=" * 70)
        print("✓ Export completed successfully!")
        print(f"✓ All files saved in: {os.path.abspath(OUTPUT_DIR)}")
        print("=" * 70)
        print(f"\n📊 EXPORT SUMMARY:")
        print(f"   • Collections exported: {successful_collections}/{len(COLLECTIONS)}")
        print(f"   • Total documents: {total_docs}")
        print(f"   • Event registration files: {total_events}")
        print(f"   • Total Excel files created: {successful_collections + total_events + 1}")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n✗ Fatal error: {e}")
        print("=" * 70)
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())
